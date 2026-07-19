"""Formula-safe ``.xlsx`` front-end for the approved test-data workbook (plan §13.1).

The workbook is a versioned test-data *contract*, so it is opened with
``data_only=False``: Stage 4 must inspect formula text and reject it rather than
silently trusting Excel's last cached value. The reader converts the workbook to
the same normalized mapping the JSON front-end produces and hands it to
``read_document`` for the existing typed validation.

Convention (plan §13.1):
  * ``Manifest`` — a two-column key/value metadata sheet.
  * one sheet per ``RecordType`` — ``record_id``/``record_status``/``name``/
    ``owner``/``valid_from_revision``/``natural_key`` envelope columns plus
    ``payload.<field>`` typed-payload columns.
  * ``Bindings`` — scenario/profile/variant-to-record bindings.

Any formula cell in a data region raises ``STRUCT_FORMULA_CELL``. Cached values
are never accepted as constants.
"""

from __future__ import annotations

import hashlib
from datetime import date, datetime
from pathlib import Path
from typing import Any

# Envelope columns whose text is a comma-separated list of IDs/tags.
_RECORD_LIST_FIELDS = frozenset({"applicable_profile_ids", "tags"})
_BINDING_LIST_FIELDS = frozenset(
    {
        "test_vector_ids",
        "oracle_ids",
        "fault_profile_ids",
        "safety_rule_ids",
        "applicable_profile_ids",
    }
)
_BOOL_FIELDS = frozenset({"enabled"})
_RESERVED_SHEETS = frozenset({"Manifest", "Bindings"})
_REQUIRED_RECORD_COLUMNS = frozenset(
    {"record_id", "record_status", "name", "owner", "valid_from_revision", "natural_key"}
)
_REQUIRED_BINDING_COLUMNS = frozenset(
    {
        "binding_id",
        "scenario_id",
        "execution_profile_id",
        "fixture_id",
        "oracle_ids",
        "cleanup_id",
        "approval_status",
    }
)

# Record sheet names recognized by the ingestion vocabulary (domain/test_data_schemas).
_RECORD_TYPES = frozenset(
    {
        "ExecutionProfile",
        "Resource",
        "Endpoint",
        "Identity",
        "InterfaceProfile",
        "DataField",
        "BitField",
        "Capability",
        "LifecycleOperation",
        "Fixture",
        "Action",
        "ActionStep",
        "Oracle",
        "TestVector",
        "TestVectorValue",
        "TimingPolicy",
        "FaultProfile",
        "Cleanup",
        "Decision",
        "SafetyRule",
        "Dependency",
    }
)


class XlsxFormulaCellError(ValueError):
    """A data-region cell contains a formula (STRUCT_FORMULA_CELL, plan §13.1)."""

    __test__ = False

    def __init__(self, *, sheet: str, cell: str, formula: str) -> None:
        self.sheet = sheet
        self.cell = cell
        self.formula = formula
        super().__init__(
            f"STRUCT_FORMULA_CELL: formula in {sheet}!{cell} ('{formula}'); "
            "formulas are rejected, cached values are never trusted as constants"
        )


def read_workbook(path: Path) -> dict[str, Any]:
    """Return a normalized ``{manifest, records, bindings, warnings}`` mapping.

    Raises ``XlsxFormulaCellError`` on the first formula found in any data region.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=False,
        keep_vba=False,
        keep_links=False,
    )
    try:
        sheet_names = list(workbook.sheetnames)
        if "Manifest" not in sheet_names:
            raise ValueError("workbook is missing the required 'Manifest' sheet")

        manifest = _read_manifest(workbook["Manifest"])
        # The file bytes, not operator-entered metadata, are authoritative.
        manifest["workbook_checksum"] = "sha256:" + hashlib.sha256(path.read_bytes()).hexdigest()
        records: list[dict[str, Any]] = []
        bindings: list[dict[str, Any]] = []
        warnings: list[str] = []

        for name in sheet_names:
            if name == "Manifest":
                continue
            if name == "Bindings":
                bindings = _read_table(
                    workbook[name],
                    name,
                    _BINDING_LIST_FIELDS,
                    with_provenance=False,
                    required_columns=_REQUIRED_BINDING_COLUMNS,
                )
                continue
            if name in _RECORD_TYPES:
                for row in _read_table(
                    workbook[name],
                    name,
                    _RECORD_LIST_FIELDS,
                    with_provenance=True,
                    required_columns=_REQUIRED_RECORD_COLUMNS,
                ):
                    records.append({**row, "record_type": name})
                continue
            warnings.append(f"unknown sheet '{name}' ignored")

        return {
            "manifest": manifest,
            "records": records,
            "bindings": bindings,
            "warnings": warnings,
        }
    finally:
        workbook.close()


def _read_manifest(sheet: Any) -> dict[str, Any]:
    manifest: dict[str, Any] = {}
    for row_index, row in enumerate(sheet.iter_rows(), start=1):
        cells = list(row)
        if not cells or cells[0].value is None:
            continue
        _reject_formula(sheet.title, cells[0])
        key = str(cells[0].value).strip()
        value: Any = None
        if len(cells) > 1:
            _reject_formula(sheet.title, cells[1])
            value = _coerce_scalar(cells[1].value)
        if row_index == 1 and key.lower() in {"key", "field"}:
            continue  # optional header row
        if key:
            manifest[key] = value
    return manifest


def _read_table(
    sheet: Any,
    sheet_name: str,
    list_fields: frozenset[str],
    *,
    with_provenance: bool,
    required_columns: frozenset[str],
) -> list[dict[str, Any]]:
    rows_iter = sheet.iter_rows()
    header_cells = next(rows_iter, None)
    if header_cells is None:
        return []
    headers: list[str] = []
    for cell in header_cells:
        _reject_formula(sheet_name, cell)
        headers.append(str(cell.value).strip() if cell.value is not None else "")
    populated = [header for header in headers if header]
    duplicates = sorted({header for header in populated if populated.count(header) > 1})
    if duplicates:
        raise ValueError(f"sheet '{sheet_name}' has duplicate columns: {duplicates}")
    missing = sorted(required_columns - set(populated))
    if missing:
        raise ValueError(f"sheet '{sheet_name}' is missing required columns: {missing}")

    table: list[dict[str, Any]] = []
    for excel_row, row in enumerate(rows_iter, start=2):
        values = list(row)
        if all(cell.value is None for cell in values):
            continue
        record: dict[str, Any] = (
            {"source_sheet": sheet_name, "source_row": excel_row} if with_provenance else {}
        )
        payload: dict[str, Any] = {}
        for header, cell in zip(headers, values, strict=False):
            _reject_formula(sheet_name, cell)
            if not header:
                continue
            raw = _coerce_scalar(cell.value)
            if header.startswith("payload."):
                payload[header[len("payload.") :]] = raw
            elif header in list_fields:
                record[header] = _split_list(raw)
            elif header in _BOOL_FIELDS:
                record[header] = _coerce_bool(raw)
            else:
                record[header] = raw
        if payload:
            record["payload"] = payload
        table.append(record)
    return table


def _reject_formula(sheet_name: str, cell: Any) -> None:
    # In read_only + data_only=False mode a formula cell reports data_type 'f'.
    if getattr(cell, "data_type", None) == "f" or (
        isinstance(cell.value, str) and cell.value.startswith("=")
    ):
        column = getattr(cell, "column_letter", "?")
        row = getattr(cell, "row", "?")
        raise XlsxFormulaCellError(sheet=sheet_name, cell=f"{column}{row}", formula=str(cell.value))


def _coerce_scalar(value: Any) -> Any:
    if isinstance(value, datetime | date):
        return value.isoformat()
    return value


def _coerce_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {"true", "1", "yes", "y"}:
        return True
    if normalized in {"false", "0", "no", "n"}:
        return False
    raise ValueError(f"invalid Boolean cell value: {value!r}")


def _split_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [part.strip() for part in str(value).split(",") if part.strip()]


__all__ = ["XlsxFormulaCellError", "read_workbook"]

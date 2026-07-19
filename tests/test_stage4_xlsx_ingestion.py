"""P1 tests: formula-safe ``.xlsx`` ingestion (plan §13.1).

Real workbooks are built with openpyxl so the ``data_only=False`` formula
rejection is exercised end-to-end. Exact typed values, bindings, secret
references, and unknown-sheet warnings are all asserted.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from multi_agentic_graph_rag.services.test_data_document_reader import (
    TestDataDocumentError,
    read_document,
)


def _base_workbook() -> Workbook:
    wb = Workbook()
    manifest = wb.active
    manifest.title = "Manifest"
    manifest.append(["key", "value"])
    manifest.append(["project", "Demo Project"])
    manifest.append(["schema_version", "1.0"])
    manifest.append(["workbook_checksum", "sha256:book"])
    manifest.append(["decision_revision", "r1"])

    oracle = wb.create_sheet("Oracle")
    oracle.append(
        [
            "record_id",
            "record_status",
            "name",
            "owner",
            "valid_from_revision",
            "natural_key",
            "payload.threshold",
            "payload.secret_ref",
        ]
    )
    oracle.append(
        ["ORC-1", "APPROVED", "Temp oracle", "qa", "r1", "temp-threshold", 42, "secret://db/pw"]
    )

    bindings = wb.create_sheet("Bindings")
    bindings.append(
        [
            "binding_id",
            "scenario_id",
            "execution_profile_id",
            "fixture_id",
            "oracle_ids",
            "cleanup_id",
            "approval_status",
        ]
    )
    bindings.append(["BND-1", "TS-1", "EP-DEFAULT", "FX-1", "ORC-1", "CLN-1", "APPROVED"])
    return wb


def _write(wb: Workbook, tmp_path: Path) -> Path:
    path = tmp_path / "test-data.xlsx"
    wb.save(path)
    return path


def test_reads_typed_records_and_bindings(tmp_path: Path) -> None:
    doc = read_document(_write(_base_workbook(), tmp_path))
    assert doc.project == "Demo Project"
    assert [issue.severity for issue in doc.structural_issues] == []

    assert len(doc.records) == 1
    record = doc.records[0]
    assert record.record_type == "Oracle"
    assert record.record_id == "ORC-1"
    # Exact typed value is preserved, not coerced to text.
    assert record.payload["threshold"] == 42
    assert isinstance(record.payload["threshold"], int)
    # Secret stays an unresolved reference (redaction happens downstream).
    assert record.payload["secret_ref"] == "secret://db/pw"
    assert record.source_sheet == "Oracle"
    assert record.source_row == 2

    assert len(doc.bindings) == 1
    assert doc.bindings[0].oracle_ids == ["ORC-1"]


def test_formula_cell_is_rejected(tmp_path: Path) -> None:
    wb = _base_workbook()
    # Inject a formula into the payload data region.
    wb["Oracle"].cell(row=2, column=7).value = "=1+1"
    with pytest.raises(TestDataDocumentError, match="STRUCT_FORMULA_CELL"):
        read_document(_write(wb, tmp_path))


def test_unknown_sheet_warns_but_parses(tmp_path: Path) -> None:
    wb = _base_workbook()
    extra = wb.create_sheet("Scratchpad")
    extra.append(["anything"])
    extra.append(["ignored"])
    doc = read_document(_write(wb, tmp_path))
    codes = {issue.issue_code for issue in doc.structural_issues}
    assert "STRUCT_UNKNOWN_SHEET" in codes
    assert len(doc.records) == 1  # unknown sheet did not become a record


def test_missing_manifest_sheet_is_error(tmp_path: Path) -> None:
    wb = Workbook()
    wb.active.title = "Oracle"
    with pytest.raises(TestDataDocumentError):
        read_document(_write(wb, tmp_path))
